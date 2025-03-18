import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import json
import re
from pathlib import Path
import pandas as pd
from keyword_manager import KeywordManager
from keyword_dialog import KeywordDialog
from tkinterdnd2 import TkinterDnD, DND_FILES
import os
import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import time
import requests
from packaging import version

class JsonExtractorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("漏扫数据分析工具 v1.0.1 - By Felix")
        self.root.geometry("800x600")
        
        # 创建菜单栏
        self.menubar = tk.Menu(self.root)
        self.root.config(menu=self.menubar)
        
        # 创建帮助菜单
        self.help_menu = tk.Menu(self.menubar, tearoff=0)
        self.menubar.add_cascade(label="帮助", menu=self.help_menu)
        self.help_menu.add_command(label="关于", command=self.show_about)
        self.help_menu.add_command(label="检查更新", command=self.check_update)
        
        # 创建标签页
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 创建漏洞分类标签页
        self.vuln_class_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.vuln_class_frame, text="漏洞类型分类")
        
        # 创建IP提取标签页
        self.ip_extract_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.ip_extract_frame, text="漏洞明细表格IP提取")
        
        # 创建漏扫结果提取标签页
        self.vuln_export_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.vuln_export_frame, text="漏扫结果提取明细至Excel")
        
        # 在漏洞分类标签页中添加组件
        # 输入文件框
        self.input_frame = tk.LabelFrame(self.vuln_class_frame, text="输入HTML文件", padx=10, pady=5)
        self.input_frame.pack(fill="x", padx=10, pady=5)
        
        self.input_path = tk.StringVar()
        self.input_entry = tk.Entry(self.input_frame, textvariable=self.input_path, width=80)
        self.input_entry.pack(side="left", padx=5)
        
        self.browse_btn = tk.Button(self.input_frame, text="浏览", command=self.browse_input)
        self.browse_btn.pack(side="left", padx=5)
        
        # 拖放提示移到输入框右边
        self.drop_label = tk.Label(self.input_frame, text="(支持拖放HTML文件)", pady=5)
        self.drop_label.pack(side="left", padx=5)
        
        # 输出文件框
        self.output_frame = tk.LabelFrame(self.vuln_class_frame, text="输出JSON文件", padx=10, pady=5)
        self.output_frame.pack(fill="x", padx=10, pady=5)
        
        self.output_path = tk.StringVar()
        self.output_entry = tk.Entry(self.output_frame, textvariable=self.output_path, width=80)
        self.output_entry.pack(side="left", padx=5)
        
        self.save_btn = tk.Button(self.output_frame, text="浏览", command=self.browse_output)
        self.save_btn.pack(side="left", padx=5)
        
        # 将提取JSON按钮移到输出文件框中
        self.extract_btn = tk.Button(self.output_frame, text="提取JSON", 
                                   command=self.extract_json)
        self.extract_btn.pack(side="left", padx=5)
        
        # 添加关键词管理器
        self.keyword_manager = KeywordManager()
        
        # 创建漏洞分类框架
        vuln_frame = tk.LabelFrame(self.vuln_class_frame, text="漏洞类型分类", padx=5, pady=5)
        vuln_frame.pack(fill="x", padx=10, pady=5)
        
        # 添加匹配模式选择
        self.match_mode = tk.StringVar(value="both")
        match_label = tk.Label(vuln_frame, text="匹配模式:")
        match_label.pack(side="left", padx=5)
        
        modes = [
            ("精准匹配", "exact"),
            ("模糊匹配", "fuzzy"),
            ("精准+模糊", "both")
        ]
        
        for text, mode in modes:
            tk.Radiobutton(vuln_frame, text=text, variable=self.match_mode, 
                          value=mode).pack(side="left", padx=5)
        
        # 添加关键词管理按钮
        self.keyword_btn = tk.Button(vuln_frame, text="关键词管理", 
                                   command=self.show_keyword_dialog)
        self.keyword_btn.pack(side="left", padx=15)
        
        # 添加导出按钮
        self.vuln_btn = tk.Button(vuln_frame, text="导出分类", 
                                 command=self.export_vuln_types)
        self.vuln_btn.pack(side="left", padx=5)
        
        # 日志框
        self.log_frame = tk.LabelFrame(self.vuln_class_frame, text="运行日志", padx=10, pady=5)
        self.log_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # 创建文本框和滚动条
        self.log_text = tk.Text(self.log_frame, height=15, width=80)
        self.scrollbar = tk.Scrollbar(self.log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=self.scrollbar.set)
        
        # 放置文本框和滚动条
        self.scrollbar.pack(side="right", fill="y")
        self.log_text.pack(side="left", fill="both", expand=True)
        
        # 状态显示
        self.status_var = tk.StringVar()
        self.status_label = tk.Label(self.vuln_class_frame, textvariable=self.status_var)
        self.status_label.pack(pady=5)
        
        # 绑定拖放事件
        self.root.drop_target_register(DND_FILES)
        self.root.dnd_bind('<<Drop>>', self.handle_drop)
        
        # 初始化IP提取标签页
        self.init_ip_extract_tab()
        
        # 初始化漏扫结果提取标签页
        self.init_lvmeng_export_tab()

    def log_message(self, message, level="INFO"):
        """添加日志消息到日志框"""
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_text.insert("end", f"[{timestamp}] [{level}] {message}\n")
        self.log_text.see("end")  # 自动滚动到最新消息
        
    def browse_input(self):
        filename = filedialog.askopenfilename(
            filetypes=[("HTML文件", "*.html"), ("所有文件", "*.*")]
        )
        if filename:
            self.input_path.set(filename)
            # 自动设置输出文件名
            output_path = Path(filename).with_suffix('.json')
            self.output_path.set(str(output_path))

    def browse_output(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON文件", "*.json"), ("所有文件", "*.*")]
        )
        if filename:
            self.output_path.set(filename)

    def browse_ip_file(self):
        filename = filedialog.askopenfilename(
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if filename:
            self.ip_file_path_var.set(filename)

    def handle_drop(self, event):
        file_path = event.data
        self.log_message(f"收到拖放文件: {file_path}")
        
        # 移除花括号并转换为Path对象
        file_path = Path(file_path.strip('{}'))
        
        # 检查文件扩展名（不区分大小写）
        if file_path.suffix.lower() in ('.html', '.htm'):
            self.input_path.set(str(file_path))
            output_path = file_path.with_suffix('.json')
            self.output_path.set(str(output_path))
            self.log_message(f"已设置输入文件: {file_path}")
            self.log_message(f"已设置输出文件: {output_path}")
        else:
            self.log_message(f"无效的文件类型: {file_path}", "ERROR")

    def handle_ip_file_drop(self, event):
        file_path = event.data
        self.log_message(f"收到拖放文件: {file_path}")
        
        # 移除花括号并转换为Path对象
        file_path = Path(file_path.strip('{}'))
        
        # 检查文件扩展名（不区分大小写）
        if file_path.suffix.lower() == '.xlsx':
            self.ip_file_path_var.set(str(file_path))
            self.log_message(f"已设置IP提取Excel文件: {file_path}")
        else:
            self.log_message(f"无效的文件类型: {file_path}", "ERROR")

    def extract_json(self):
        input_file = self.input_path.get()
        output_file = self.output_path.get()
        
        if not input_file or not output_file:
            self.log_message("请选择输入和输出文件！", "ERROR")
            return
            
        try:
            # 尝试多种编码格式
            encodings = ['utf-8', 'gbk', 'gb2312', 'iso-8859-1']
            file_content = None
            
            for encoding in encodings:
                try:
                    with open(input_file, encoding=encoding) as f:
                        file_content = f.read()
                    self.log_message(f"成功使用 {encoding} 编码读取文件")
                    break
                except UnicodeDecodeError:
                    continue
            
            if file_content is None:
                self.log_message(f"无法读取文件，已尝试以下编码: {', '.join(encodings)}", "ERROR")
                return
            
            self.log_message(f"正在处理文件: {input_file}")
            pat_list = re.findall(r'<script>window.data = (.*?);</script>', file_content)
            
            if not pat_list:
                self.log_message("未在HTML文件中找到匹配的JSON数据！", "ERROR")
                return
                
            data_json = json.loads(pat_list[0])
            
            with open(output_file, 'w', encoding='utf8') as f:
                json.dump(data_json, f, ensure_ascii=False, indent=4)
                
            success_msg = f"JSON数据已成功保存到: {output_file}"
            self.log_message(success_msg, "SUCCESS")
            self.status_var.set("JSON提取成功！")
            
        except Exception as e:
            error_msg = f"处理过程中出现错误：{str(e)}"
            self.log_message(error_msg, "ERROR")
            self.status_var.set(f"错误: {str(e)}")

    def show_keyword_dialog(self):
        KeywordDialog(self.root, self.keyword_manager)
    
    def export_vuln_types(self):
        try:
            # 先让用户选择保存位置
            output_excel = filedialog.asksaveasfilename(
                title="保存漏洞分类Excel",
                defaultextension=".xlsx",
                initialfile="漏洞类型分类.xlsx",
                filetypes=[
                    ("Excel文件", "*.xlsx"),
                    ("所有文件", "*.*")
                ]
            )
            
            if not output_excel:  # 用户取消选择
                return
            
            # 读取JSON文件
            with open(self.output_path.get(), 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 获取漏洞列表
            vuln_list = data['categories'][3]['children'][0]['data']['vulns_info']['vuln_distribution']['vuln_list']
            
            # 准备Excel数据
            excel_data = []
            for i, vuln in enumerate(vuln_list, 1):
                # 处理描述和解决方案
                description = '\n'.join(filter(None, vuln.get('i18n_description', [])))
                solution = '\n'.join(filter(None, vuln.get('i18n_solution', [])))
                
                # 获取漏洞等级中文
                level_map = {'high': '高危', 'middle': '中危', 'low': '低危'}
                level = level_map.get(vuln.get('vuln_level', ''), '未知')
                
                # 获取漏洞类型
                vuln_type = self.keyword_manager.get_type(
                    vuln.get('i18n_name', ''), 
                    self.match_mode.get()
                )
                
                excel_data.append({
                    '序号': i,
                    '漏洞名称': vuln.get('i18n_name', ''),
                    '类型': vuln_type,
                    '漏洞等级': level,
                    '影响主机个数': vuln.get('vuln_count', 0),
                    '受影响主机': vuln.get('target', ''),
                    '详细描述': description,
                    '解决办法': solution
                })
            
            # 创建DataFrame并导出到Excel
            df = pd.DataFrame(excel_data)
            df.to_excel(output_excel, index=False)
            
            self.log_message(f"漏洞类型分类已导出到: {output_excel}", "SUCCESS")
            
        except Exception as e:
            self.log_message(f"导出漏洞类型分类时出错: {str(e)}", "ERROR")

    def init_ip_extract_tab(self):
        # 添加说明文字
        description = "本功能旨在漏洞明细Excel表格按漏洞等级、漏洞类型进行提取IP地址，支持IP去重，请先将漏洞明细Excel表格进行漏洞分类后再使用本功能！"
        desc_label = tk.Label(self.ip_extract_frame, text=description, wraplength=700, justify="left")
        desc_label.pack(fill="x", padx=10, pady=10)
        
        # 文件选择框架
        file_frame = tk.LabelFrame(self.ip_extract_frame, text="选择Excel文件", padx=10, pady=5)
        file_frame.pack(fill="x", padx=10, pady=5)
        
        self.ip_file_path_var = tk.StringVar()
        self.ip_file_entry = tk.Entry(file_frame, textvariable=self.ip_file_path_var, width=80)
        self.ip_file_entry.pack(side="left", padx=5)
        
        browse_btn = tk.Button(file_frame, text="浏览", command=self.browse_ip_file)
        browse_btn.pack(side="left", padx=5)
        
        # Excel类型选择框架
        type_frame = tk.LabelFrame(self.ip_extract_frame, text="Excel类型", padx=10, pady=5)
        type_frame.pack(fill="x", padx=10, pady=5)
        
        self.excel_type_var = tk.StringVar(value="complex")
        simple_radio = tk.Radiobutton(type_frame, text="简单表格(一列一个字段)", 
                                     variable=self.excel_type_var, value="simple")
        simple_radio.pack(side="left", padx=5)
        complex_radio = tk.Radiobutton(type_frame, text="复杂表格(混合布局字段)", 
                                      variable=self.excel_type_var, value="complex")
        complex_radio.pack(side="left", padx=5)
        
        # 去重选项
        self.ip_deduplicate_var = tk.BooleanVar()
        dedup_check = tk.Checkbutton(self.ip_extract_frame, text="去重IP地址", 
                                   variable=self.ip_deduplicate_var)
        dedup_check.pack(pady=10)
        
        # 提取按钮
        extract_btn = tk.Button(self.ip_extract_frame, text="提取IP地址", 
                              command=self.extract_ip_addresses)
        extract_btn.pack(pady=10)
        
        # IP提取状态标签
        self.ip_status_var = tk.StringVar()
        ip_status_label = tk.Label(self.ip_extract_frame, 
                                 textvariable=self.ip_status_var, fg="green")
        ip_status_label.pack(pady=5)
        
        # 文件拖放支持
        self.ip_file_entry.drop_target_register(DND_FILES)
        self.ip_file_entry.dnd_bind('<<Drop>>', self.handle_ip_file_drop)

    def extract_ip_addresses(self):
        file_path = self.ip_file_path_var.get()
        if not file_path:
            messagebox.showwarning("警告", "请选择一个有效的Excel文件。")
            return

        # 让用户选择导出目录
        output_dir = filedialog.askdirectory(title="选择IP地址文件保存目录")
        if not output_dir:  # 用户取消选择
            return

        try:
            df = pd.read_excel(file_path)
            
            # 初始化文件句柄字典和IP集合字典
            file_handles = {}
            ip_sets = {}
            
            if self.excel_type_var.get() == "simple":
                # 处理简单表格格式
                for _, row in df.iterrows():
                    risk_level = str(row['漏洞等级']).strip()
                    category = str(row['类型']).strip()
                    ip_addresses = str(row['受影响主机']).strip()
                    
                    if ip_addresses and ip_addresses != 'nan':
                        # 构建完整的文件路径
                        filename = Path(output_dir) / f"{risk_level}-{category}.txt"
                        
                        # 如果文件句柄不存在，则创建新的文件句柄和IP集合
                        if filename not in file_handles:
                            file_handles[filename] = open(filename, 'w', encoding='utf-8')
                            if self.ip_deduplicate_var.get():
                                ip_sets[filename] = set()
                        
                        # 如果有多个IP，使用";"分隔
                        for ip in ip_addresses.split(';'):
                            ip = ip.strip()
                            if ip:
                                if self.ip_deduplicate_var.get():
                                    if ip not in ip_sets[filename]:
                                        file_handles[filename].write(ip + '\n')
                                        ip_sets[filename].add(ip)
                                else:
                                    file_handles[filename].write(ip + '\n')
            else:
                # 处理复杂表格格式
                for i in range(1, len(df), 4):
                    if i + 1 < len(df):
                        risk_level = str(df.iloc[i-1, 2]).strip()
                        category = str(df.iloc[i-1, 4]).strip()
                        
                        # 构建完整的文件路径
                        filename = Path(output_dir) / f"{risk_level}-{category}.txt"
                        
                        # 如果文件句柄不存在，则创建新的文件句柄和IP集合
                        if filename not in file_handles:
                            file_handles[filename] = open(filename, 'w', encoding='utf-8')
                            if self.ip_deduplicate_var.get():
                                ip_sets[filename] = set()
                        
                        ip_addresses = str(df.iloc[i, 3]).strip()
                        if ip_addresses and ip_addresses != 'nan':
                            # 如果有多个IP，使用";"分隔
                            for ip in ip_addresses.split(';'):
                                ip = ip.strip()
                                if ip:
                                    if self.ip_deduplicate_var.get():
                                        if ip not in ip_sets[filename]:
                                            file_handles[filename].write(ip + '\n')
                                            ip_sets[filename].add(ip)
                                    else:
                                        file_handles[filename].write(ip + '\n')
            
            # 确保所有文件都被正确关闭
            for fh in file_handles.values():
                fh.close()
            
            self.ip_status_var.set(f"IP地址已成功提取并保存到目录: {output_dir}")
            messagebox.showinfo("完成", f"IP地址已成功提取并保存到目录: {output_dir}")
        except Exception as e:
            self.ip_status_var.set(f"发生错误: {str(e)}")
            messagebox.showerror("错误", f"发生错误: {str(e)}")

    def init_lvmeng_export_tab(self):
        """初始化漏扫结果提取明细至Excel标签页"""
        # 添加说明文字
        description = "本功能用于将绿盟漏扫系统生成的漏洞扫描结果HTML源文件提取至Excel文件，支持新旧版绿盟的漏洞扫描结果的解析和导出，并提供不同的导出样式。"
        desc_label = tk.Label(self.vuln_export_frame, text=description, wraplength=700, justify="left")
        desc_label.pack(fill="x", padx=10, pady=10)
        
        # 输入文件框架
        input_frame = tk.LabelFrame(self.vuln_export_frame, text="输入HTML文件", padx=10, pady=5)
        input_frame.pack(fill="x", padx=10, pady=5)
        
        self.lvmeng_input_path = tk.StringVar()
        self.lvmeng_input_entry = tk.Entry(input_frame, textvariable=self.lvmeng_input_path, width=80)
        self.lvmeng_input_entry.pack(side="left", padx=5)
        
        browse_btn = tk.Button(input_frame, text="浏览", command=self.browse_lvmeng_input)
        browse_btn.pack(side="left", padx=5)
        
        # 拖放提示
        drop_label = tk.Label(input_frame, text="(支持拖放HTML文件)", pady=5)
        drop_label.pack(side="left", padx=5)
        
        # 版本选择框架
        version_frame = tk.LabelFrame(self.vuln_export_frame, text="绿盟版本", padx=10, pady=5)
        version_frame.pack(fill="x", padx=10, pady=5)
        
        self.lvmeng_version = tk.StringVar(value="new")
        new_radio = tk.Radiobutton(version_frame, text="新版绿盟", 
                                  variable=self.lvmeng_version, value="new")
        new_radio.pack(side="left", padx=5)
        old_radio = tk.Radiobutton(version_frame, text="旧版绿盟", 
                                  variable=self.lvmeng_version, value="old")
        old_radio.pack(side="left", padx=5)
        
        # 样式选择框架
        style_frame = tk.LabelFrame(self.vuln_export_frame, text="导出样式", padx=10, pady=5)
        style_frame.pack(fill="x", padx=10, pady=5)
        
        self.lvmeng_style = tk.StringVar(value="style1")
        style1_radio = tk.Radiobutton(style_frame, text="样式一(简单表格)", 
                                    variable=self.lvmeng_style, value="style1")
        style1_radio.pack(side="left", padx=5)
        style2_radio = tk.Radiobutton(style_frame, text="样式二(复杂表格)", 
                                    variable=self.lvmeng_style, value="style2")
        style2_radio.pack(side="left", padx=5)
        
        # 输出文件框架
        output_frame = tk.LabelFrame(self.vuln_export_frame, text="输出Excel文件", padx=10, pady=5)
        output_frame.pack(fill="x", padx=10, pady=5)
        
        self.lvmeng_output_path = tk.StringVar()
        self.lvmeng_output_entry = tk.Entry(output_frame, textvariable=self.lvmeng_output_path, width=80)
        self.lvmeng_output_entry.pack(side="left", padx=5)
        
        browse_output_btn = tk.Button(output_frame, text="浏览", command=self.browse_lvmeng_output)
        browse_output_btn.pack(side="left", padx=5)
        
        # 导出按钮
        export_btn = tk.Button(self.vuln_export_frame, text="导出Excel", 
                              command=self.export_lvmeng_to_excel)
        export_btn.pack(pady=10)
        
        # 漏扫提取状态标签
        self.lvmeng_status_var = tk.StringVar()
        lvmeng_status_label = tk.Label(self.vuln_export_frame, 
                                   textvariable=self.lvmeng_status_var, fg="green")
        lvmeng_status_label.pack(pady=5)
        
        # 日志框
        log_frame = tk.LabelFrame(self.vuln_export_frame, text="运行日志", padx=10, pady=5)
        log_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # 创建文本框和滚动条
        self.lvmeng_log_text = tk.Text(log_frame, height=12, width=80)
        scrollbar = tk.Scrollbar(log_frame, orient="vertical", command=self.lvmeng_log_text.yview)
        self.lvmeng_log_text.configure(yscrollcommand=scrollbar.set)
        
        # 放置文本框和滚动条
        scrollbar.pack(side="right", fill="y")
        self.lvmeng_log_text.pack(side="left", fill="both", expand=True)
        
        # 文件拖放支持
        self.lvmeng_input_entry.drop_target_register(DND_FILES)
        self.lvmeng_input_entry.dnd_bind('<<Drop>>', self.handle_lvmeng_file_drop)

    def browse_lvmeng_input(self):
        """浏览并选择HTML输入文件"""
        filename = filedialog.askopenfilename(
            filetypes=[("HTML文件", "*.html"), ("所有文件", "*.*")]
        )
        if filename:
            self.lvmeng_input_path.set(filename)
            # 自动设置输出文件名
            output_path = Path(filename).with_suffix('.xlsx')
            self.lvmeng_output_path.set(str(output_path))
            self.lvmeng_log("已选择输入文件: " + filename)

    def browse_lvmeng_output(self):
        """浏览并选择Excel输出文件"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if filename:
            self.lvmeng_output_path.set(filename)
            self.lvmeng_log("已选择输出文件: " + filename)

    def handle_lvmeng_file_drop(self, event):
        """处理拖放HTML文件"""
        file_path = event.data
        self.lvmeng_log(f"收到拖放文件: {file_path}")
        
        # 移除花括号并转换为Path对象
        file_path = Path(file_path.strip('{}'))
        
        # 检查文件扩展名（不区分大小写）
        if file_path.suffix.lower() in ('.html', '.htm'):
            self.lvmeng_input_path.set(str(file_path))
            output_path = file_path.with_suffix('.xlsx')
            self.lvmeng_output_path.set(str(output_path))
            self.lvmeng_log(f"已设置输入文件: {file_path}")
            self.lvmeng_log(f"已设置输出文件: {output_path}")
        else:
            self.lvmeng_log(f"无效的文件类型: {file_path}", "ERROR")

    def lvmeng_log(self, message, level="INFO"):
        """添加日志消息到漏扫提取日志框"""
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.lvmeng_log_text.insert("end", f"[{timestamp}] [{level}] {message}\n")
        self.lvmeng_log_text.see("end")  # 自动滚动到最新消息

    def export_lvmeng_to_excel(self):
        """导出漏扫结果到Excel"""
        input_file = self.lvmeng_input_path.get()
        output_file = self.lvmeng_output_path.get()
        version = self.lvmeng_version.get()
        style = self.lvmeng_style.get()
        
        if not input_file or not output_file:
            messagebox.showwarning("警告", "请选择输入和输出文件。")
            self.lvmeng_log("请选择输入和输出文件！", "ERROR")
            return
        
        if not Path(input_file).exists():
            messagebox.showwarning("警告", "输入文件不存在。")
            self.lvmeng_log("输入文件不存在！", "ERROR")
            return
        
        try:
            self.lvmeng_log(f"开始处理文件: {input_file}")
            self.lvmeng_log(f"导出版本: {'新版绿盟' if version == 'new' else '旧版绿盟'}")
            self.lvmeng_log(f"导出样式: {'样式一' if style == 'style1' else '样式二'}")
            
            # 根据版本和样式选择处理方法
            if version == "new":
                self.process_new_lvmeng(input_file, output_file, style)
            else:
                self.process_old_lvmeng(input_file, output_file, style)
            
            self.lvmeng_status_var.set(f"漏洞数据已成功导出到: {output_file}")
            messagebox.showinfo("完成", f"漏洞数据已成功导出到: {output_file}")
        except Exception as e:
            error_msg = f"导出过程中出现错误: {str(e)}"
            self.lvmeng_log(error_msg, "ERROR")
            self.lvmeng_status_var.set("导出失败")
            messagebox.showerror("错误", error_msg)

    def process_new_lvmeng(self, input_file, output_file, style):
        """处理新版绿盟漏扫结果并导出到Excel"""
        try:
            # 读取HTML文件
            with open(input_file, 'r', encoding='utf-8') as f:
                file_content = f.read()
            
            # 提取JSON数据
            self.lvmeng_log("正在提取JSON数据...")
            pat_list = re.findall(r'<script>window.data = (.*?);</script>', file_content)
            if not pat_list:
                raise Exception("未在HTML文件中找到匹配的JSON数据")
            
            data_json = json.loads(pat_list[0])
            
            # 获取漏洞列表
            vuln_list = data_json["categories"][3]["children"][0]["data"]["vulns_info"]["vuln_distribution"]["vuln_list"]
            self.lvmeng_log(f"成功提取到 {len(vuln_list)} 个漏洞")
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "漏洞信息"
            
            # 定义样式
            # 定义对齐方式（水平居中，垂直居中）
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # 定义对齐方式（水平左对齐，垂直居中）
            horLeft_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            # 单独定义字体加粗
            font_bold = Font(bold=True)
            # 单独定义颜色字体
            font_red = Font(bold=False, color='E42B00')
            font_orange = Font(bold=False, color='AF6100')
            font_gray = Font(bold=False, color='737373')
            # 定义边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if style == 'style1':
                # 样式一：简单表格
                self.lvmeng_log("使用样式一(简单表格)导出...")
                
                # 添加表头
                headers = ['序号', '漏洞名称', '漏洞等级', '影响主机个数', '受影响主机', '详细描述', '解决办法']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.alignment = alignment
                    cell.font = font_bold
                    cell.border = thin_border
                
                # 设置列宽
                ws.column_dimensions['A'].width = 7
                ws.column_dimensions['B'].width = 45
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 40
                ws.column_dimensions['F'].width = 50
                ws.column_dimensions['G'].width = 50
                
                # 添加数据
                for idx, vuln in enumerate(vuln_list, 1):
                    # 获取漏洞等级中文
                    level_map = {'high': '高危', 'middle': '中危', 'low': '低危'}
                    level = level_map.get(vuln.get('vuln_level', ''), '未知')
                    
                    # 处理描述和解决方案
                    description = '\n'.join(filter(None, vuln.get('i18n_description', [])))
                    solution = '\n'.join(filter(None, vuln.get('i18n_solution', [])))
                    
                    # 添加行数据
                    row_data = [
                        idx,
                        vuln.get('i18n_name', ''),
                        level,
                        vuln.get('vuln_count', 0),
                        vuln.get('target', ''),
                        description,
                        solution
                    ]
                    
                    row_idx = idx + 1
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = thin_border
                        
                        # 设置对齐方式
                        if col_idx in [1, 3, 4]:  # 序号、漏洞等级、影响主机个数居中
                            cell.alignment = alignment
                        else:  # 其他左对齐
                            cell.alignment = horLeft_alignment
                        
                        # 设置漏洞等级颜色
                        if col_idx in [2, 3]:  # 漏洞名称和等级
                            if level == '高危':
                                cell.font = font_red
                            elif level == '中危':
                                cell.font = font_orange
                            elif level == '低危':
                                cell.font = font_gray
            
            else:
                # 样式二：复杂表格
                self.lvmeng_log("使用样式二(复杂表格)导出...")
                
                # 添加表头
                headers = ['序号', '漏洞名称', '漏洞等级', '影响主机个数']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.alignment = alignment
                    cell.font = font_bold
                    cell.border = thin_border
                
                # 设置列宽
                ws.column_dimensions['A'].width = 7
                ws.column_dimensions['B'].width = 45
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 50
                
                current_row = 2
                for idx, vuln in enumerate(vuln_list, 1):
                    # 获取漏洞等级中文
                    level_map = {'high': '高危', 'middle': '中危', 'low': '低危'}
                    level = level_map.get(vuln.get('vuln_level', ''), '未知')
                    
                    # 处理描述和解决方案
                    description = '\n'.join(filter(None, vuln.get('i18n_description', [])))
                    solution = '\n'.join(filter(None, vuln.get('i18n_solution', [])))
                    
                    # 添加主行数据
                    row_data = [
                        idx,
                        vuln.get('i18n_name', ''),
                        level,
                        vuln.get('vuln_count', 0)
                    ]
                    
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.border = thin_border
                        
                        # 设置对齐方式
                        if col_idx == 2:  # 漏洞名称左对齐
                            cell.alignment = horLeft_alignment
                        else:  # 其他居中
                            cell.alignment = alignment
                        
                        # 设置漏洞等级颜色
                        if col_idx in [2, 3]:  # 漏洞名称和等级
                            if level == '高危':
                                cell.font = font_red
                            elif level == '中危':
                                cell.font = font_orange
                            elif level == '低危':
                                cell.font = font_gray
                    
                    # 添加详细信息行
                    details = [
                        ('受影响主机', vuln.get('target', '')),
                        ('详细描述', description),
                        ('解决办法', solution)
                    ]
                    
                    # 计算实际需要的行数
                    total_rows = len(details)
                    
                    # 合并第一列的单元格
                    if total_rows > 0:
                        ws.merge_cells(f'A{current_row+1}:A{current_row+total_rows}')
                        merged_cell = ws.cell(row=current_row+1, column=1)
                        merged_cell.border = thin_border
                    
                    # 添加所有详细信息
                    for i, (label, value) in enumerate(details):
                        row_num = current_row + i + 1
                        
                        # 添加标签（第二列）
                        label_cell = ws.cell(row=row_num, column=2)
                        label_cell.value = label
                        label_cell.font = font_bold
                        label_cell.border = thin_border
                        label_cell.alignment = alignment  # 标签居中对齐
                        
                        # 添加值（第三列）- 保持为空
                        value_cell = ws.cell(row=row_num, column=3)
                        value_cell.border = thin_border
                        
                        # 第四列设置值和自动换行
                        value_cell = ws.cell(row=row_num, column=4)
                        value_cell.value = value
                        value_cell.border = thin_border
                        value_cell.alignment = horLeft_alignment  # 自动换行
                    
                    current_row += total_rows + 1  # 更新行号，加1是为了下一组数据之间留空
            
            # 保存Excel文件
            wb.save(output_file)
            self.lvmeng_log(f"成功导出到Excel文件: {output_file}", "SUCCESS")
            return True
        
        except Exception as e:
            self.lvmeng_log(f"处理新版绿盟漏扫结果时出错: {str(e)}", "ERROR")
            raise

    def process_old_lvmeng(self, input_file, output_file, style):
        """处理旧版绿盟漏扫结果并导出到Excel"""
        try:
            # 读取HTML文件
            with open(input_file, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            # 使用BeautifulSoup解析HTML
            self.lvmeng_log("正在解析HTML文件...")
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # 查找漏洞分布表格
            vuln_table = soup.find('table', id='vuln_distribution')
            if not vuln_table:
                raise Exception("未找到漏洞分布表格，请确认是否为旧版绿盟漏扫结果")
            
            # 查找所有漏洞行
            vuln_rows = vuln_table.find_all('tr', class_=re.compile(r'(odd|even) vuln_(high|middle|low)'))
            self.lvmeng_log(f"成功提取到 {len(vuln_rows)} 个漏洞")
            
            # 存储漏洞数据
            vulnerability_data = []
            
            for row in vuln_rows:
                # 提取基本信息
                tds = row.find_all('td')
                number = tds[0].text.strip()
                name = row.find('span').text.strip()
                affected_hosts_count = tds[2].text.strip()
                
                # 确定漏洞等级
                if 'vuln_high' in row['class']:
                    level = '高危'
                elif 'vuln_middle' in row['class']:
                    level = '中危'
                else:
                    level = '低危'
                
                # 获取详细信息所在的下一个tr
                detail_row = row.find_next('tr', class_=re.compile(r'more hide (odd|even)'))
                
                # 初始化变量
                affected_hosts_str = ""
                description = ""
                solution = ""
                additional_fields = {
                    '威胁分值': '',
                    '危险插件': '',
                    '发现日期': '',
                    'CVE编号': '',
                    'CNNVD编号': '',
                    'CNCVE编号': '',
                    'CVSS评分': '',
                    'CNVD编号': ''
                }
                
                if detail_row:
                    # 提取受影响主机
                    affected_hosts = []
                    # 先查找所有<a>标签的主机
                    host_links = detail_row.find_all('a', href=re.compile(r'host/.*\.html'))
                    if host_links:  # 如果找到<a>标签
                        for link in host_links:
                            affected_hosts.append(link.text.strip())
                    else:  # 如果没有<a>标签，查找特定的td标签
                        # 查找class为report_table的table下的width为80%的td
                        report_table = detail_row.find('table', class_='report_table')
                        if report_table:
                            host_td = report_table.find('td', attrs={'width': '80%'})
                            if host_td:
                                # 替换所有&nbsp;为空格，然后获取文本
                                hosts_text = host_td.text.replace('&nbsp;', ' ').strip()
                                # 如果文本不为空，添加到列表
                                if hosts_text:
                                    affected_hosts.append(hosts_text)
                
                    # 将所有主机信息合并，用逗号分隔
                    affected_hosts_str = ', '.join(affected_hosts)
                    
                    # 提取详细描述
                    description_row = detail_row.find('tr', class_='even')
                    description = description_row.find('td').text.strip() if description_row else ''
                    
                    # 提取解决办法
                    solution_row = description_row.find_next('tr', class_='odd') if description_row else None
                    solution = solution_row.find('td').text.strip() if solution_row else ''
                    
                    # 查找所有可能包含新字段的行
                    info_rows = detail_row.find_all('tr', class_=re.compile(r'(odd|even)'))
                    for info_row in info_rows:
                        field_name = info_row.find('th')
                        if field_name and field_name.text.strip() in additional_fields:
                            field_value = info_row.find('td')
                            if field_value:
                                additional_fields[field_name.text.strip()] = field_value.text.strip()
                
                # 添加到漏洞数据列表
                vulnerability_data.append({
                    '序号': number,
                    '漏洞名称': name,
                    '漏洞等级': level,
                    '影响主机个数': affected_hosts_count,
                    '受影响主机': affected_hosts_str,
                    '详细描述': description,
                    '解决办法': solution,
                    **additional_fields  # 添加新字段
                })
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "漏洞信息"
            
            # 设置边框样式
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置表头字体
            header_font = Font(bold=True)
            
            # 设置不同等级的字体颜色
            high_font = Font(color='FF0000')  # 红色
            middle_font = Font(color='FFA500')  # 橙色
            low_font = Font(color='808080')  # 灰色
            
            if style == 'style1':
                # 样式一：简单表格
                self.lvmeng_log("使用样式一(简单表格)导出...")
                
                # 添加表头
                headers = ['序号', '漏洞名称', '漏洞等级', '影响主机个数', '受影响主机', '详细描述', '解决办法']
                ws.append(headers)
                
                # 设置表头格式
                for cell in ws[1]:
                    cell.font = header_font
                    cell.border = border
                
                # 添加数据并设置格式
                for row_data in vulnerability_data:
                    row = [row_data[h] for h in headers]
                    ws.append(row)
                    row_num = ws.max_row
                    
                    # 设置单元格边框和字体颜色
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_num, column=col)
                        cell.border = border
                        
                        # 设置漏洞名称和等级的字体颜色
                        if col in [2, 3]:  # 漏洞名称和漏洞等级列
                            level = row_data['漏洞等级']
                            if level == '高危':
                                cell.font = high_font
                            elif level == '中危':
                                cell.font = middle_font
                            else:
                                cell.font = low_font
            
                # 调整列宽
                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
            else:
                # 样式二：复杂表格
                self.lvmeng_log("使用样式二(复杂表格)导出...")
                
                # 设置居中对齐
                center_alignment = Alignment(horizontal='center', vertical='center')
                # 设置自动换行对齐
                wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # 添加表头
                headers = ['序号', '漏洞名称', '漏洞等级', '影响主机个数']
                ws.append(headers)
                
                # 设置表头格式
                for cell in ws[1]:
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = center_alignment
                
                # 设置固定列宽
                ws.column_dimensions['A'].width = 7  # 第一列
                ws.column_dimensions['B'].width = 45  # 第二列
                ws.column_dimensions['C'].width = 15  # 第三列
                ws.column_dimensions['D'].width = 50  # 第四列
                
                current_row = 2
                for row_data in vulnerability_data:
                    # 添加主行
                    row = [row_data['序号'], row_data['漏洞名称'], row_data['漏洞等级'], row_data['影响主机个数']]
                    for col, value in enumerate(row, 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.value = value
                        cell.border = border
                        cell.alignment = center_alignment  # 第一行所有单元格居中
                        
                        # 第一行第二个单元格设置自动换行
                        if col == 2:
                            cell.alignment = wrap_alignment
                        
                        # 设置漏洞名称和等级的字体颜色
                        if col in [2, 3]:  # 漏洞名称和漏洞等级列
                            level = row_data['漏洞等级']
                            if level == '高危':
                                cell.font = high_font
                            elif level == '中危':
                                cell.font = middle_font
                            else:
                                cell.font = low_font
                    
                    # 添加详细信息行
                    details = [
                        ('受影响主机', row_data['受影响主机']),
                        ('详细描述', row_data['详细描述']),
                        ('解决办法', row_data['解决办法'])
                    ]
                    
                    # 添加新字段（只添加有值的字段）
                    additional_fields = [
                        ('威胁分值', row_data['威胁分值']),
                        ('危险插件', row_data['危险插件']),
                        ('发现日期', row_data['发现日期']),
                        ('CVE编号', row_data['CVE编号']),
                        ('CNNVD编号', row_data['CNNVD编号']),
                        ('CNCVE编号', row_data['CNCVE编号']),
                        ('CVSS评分', row_data['CVSS评分']),
                        ('CNVD编号', row_data['CNVD编号'])
                    ]
                    
                    # 过滤掉空值的字段
                    details.extend([(label, value) for label, value in additional_fields if value])
                    
                    # 计算实际需要的行数
                    total_rows = len(details)
                    
                    # 合并第一列的单元格
                    if total_rows > 0:
                        ws.merge_cells(f'A{current_row+1}:A{current_row+total_rows}')
                        merged_cell = ws.cell(row=current_row+1, column=1)
                        merged_cell.border = border
                    
                    # 添加所有详细信息
                    for i, (label, value) in enumerate(details):
                        row_num = current_row + i + 1
                        
                        # 添加标签（第二列）
                        label_cell = ws.cell(row=row_num, column=2)
                        label_cell.value = label
                        label_cell.font = header_font
                        label_cell.border = border
                        label_cell.alignment = center_alignment  # 标签居中对齐
                        
                        # 添加值（第三列）- 保持为空
                        value_cell = ws.cell(row=row_num, column=3)
                        value_cell.border = border
                        
                        # 第四列设置值和自动换行
                        value_cell = ws.cell(row=row_num, column=4)
                        value_cell.value = value
                        value_cell.border = border
                        value_cell.alignment = wrap_alignment  # 自动换行
                    
                    current_row += total_rows + 1  # 更新行号，加1是为了下一组数据之间留空
            
            # 保存Excel文件
            wb.save(output_file)
            self.lvmeng_log(f"成功导出到Excel文件: {output_file}", "SUCCESS")
            return True
        
        except Exception as e:
            self.lvmeng_log(f"处理旧版绿盟漏扫结果时出错: {str(e)}", "ERROR")
            raise

    def show_about(self):
        about_text = "漏扫数据分析工具 v1.0.1\n\n"
        about_text += "作者: Felix\n"
        about_text += "开源地址: https://github.com/Felix-sec/VulnDataAnalyzer"
        messagebox.showinfo("关于", about_text)
    
    def check_update(self):
        try:
            # 获取Github最新release版本
            api_url = "https://api.github.com/repos/Felix-sec/VulnDataAnalyzer/releases/latest"
            response = requests.get(api_url)
            response.raise_for_status()
            latest_version = response.json()["tag_name"].strip("v")
            current_version = "1.0.1"
            
            if version.parse(latest_version) > version.parse(current_version):
                update_text = f"发现新版本: v{latest_version}\n当前版本: v{current_version}\n\n"
                update_text += "请访问项目地址下载最新版本：\nhttps://github.com/Felix-sec/VulnDataAnalyzer/releases"
                messagebox.showinfo("检查更新", update_text)
            else:
                messagebox.showinfo("检查更新", "当前已是最新版本！")
        except Exception as e:
            messagebox.showerror("检查更新", f"检查更新失败：{str(e)}")

if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = JsonExtractorGUI(root)
    root.mainloop()